from __future__ import annotations

from pathlib import Path
from typing import Sequence, Mapping

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
STATUS_FILLS = {
    "CONFIRMED": PatternFill("solid", fgColor="C6E0B4"),
    "WATCH": PatternFill("solid", fgColor="FFF2CC"),
    "REJECTED": PatternFill("solid", fgColor="F4CCCC"),
}


def write_result_workbook(path: Path, signals: pd.DataFrame, config_rows: Sequence[Mapping[str, object]], calibration_summary: pd.DataFrame | None = None) -> Path:
    path = Path(path); path.parent.mkdir(parents=True, exist_ok=True)
    ordered = signals.sort_values(["Status", "Score"], ascending=[True, False]) if not signals.empty else signals
    sheets = {
        "candidates": ordered[ordered["Status"].isin(["CONFIRMED", "WATCH"])] if not ordered.empty else ordered,
        "confirmed": ordered[ordered["Status"] == "CONFIRMED"] if not ordered.empty else ordered,
        "watch": ordered[ordered["Status"] == "WATCH"] if not ordered.empty else ordered,
        "all_results": ordered,
        "config": pd.DataFrame(config_rows),
    }
    if calibration_summary is not None and not calibration_summary.empty:
        sheets["probability_calibration"] = calibration_summary
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        for c in ws[1]:
            c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = Alignment(horizontal="center")
        headers = {c.value:i for i,c in enumerate(ws[1],1)}
        status_col = headers.get("Status")
        if status_col:
            for r in range(2, ws.max_row+1):
                v = ws.cell(r,status_col).value
                if v in STATUS_FILLS:
                    ws.cell(r,status_col).fill = STATUS_FILLS[v]; ws.cell(r,status_col).font=Font(bold=True)
        for idx, cells in enumerate(ws.columns,1):
            width = min(55, max(10, max((len(str(c.value)) if c.value is not None else 0 for c in cells), default=0)+2))
            ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)
    return path
