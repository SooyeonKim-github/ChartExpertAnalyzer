from __future__ import annotations

from pathlib import Path
from typing import Mapping, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
STATUS_FILLS = {
    "CONFIRMED": PatternFill("solid", fgColor="C6E0B4"),
    "WATCH": PatternFill("solid", fgColor="FFF2CC"),
    "REJECTED": PatternFill("solid", fgColor="F4CCCC"),
}


def _performance_summary(frame: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    if frame.empty or "D+20_Close_Return_Pct" not in frame.columns:
        return pd.DataFrame()

    src = frame.copy()
    src["D+20_Close_Return_Pct"] = pd.to_numeric(src["D+20_Close_Return_Pct"], errors="coerce")
    src["MFE_20D_Pct"] = pd.to_numeric(src.get("MFE_20D_Pct"), errors="coerce")
    src["MAE_20D_Pct"] = pd.to_numeric(src.get("MAE_20D_Pct"), errors="coerce")

    rows = []
    grouped = src.groupby(group_cols, dropna=False) if group_cols else [((), src)]
    for key, grp in grouped:
        valid = grp["D+20_Close_Return_Pct"].dropna()
        if valid.empty:
            continue
        keys = key if isinstance(key, tuple) else (key,)
        row = {col: value for col, value in zip(group_cols, keys)}
        row.update(
            {
                "Signal_Count": int(len(grp)),
                "Complete_20D_Count": int(len(valid)),
                "D20_Win_Rate": round(float((valid > 0).mean()), 4),
                "D20_Avg_Return_Pct": round(float(valid.mean()), 3),
                "D20_Median_Return_Pct": round(float(valid.median()), 3),
                "D20_Max_Return_Pct": round(float(valid.max()), 3),
                "D20_Min_Return_Pct": round(float(valid.min()), 3),
                "Avg_MFE_20D_Pct": round(float(grp.loc[valid.index, "MFE_20D_Pct"].mean()), 3),
                "Avg_MAE_20D_Pct": round(float(grp.loc[valid.index, "MAE_20D_Pct"].mean()), 3),
            }
        )
        rows.append(row)
    return pd.DataFrame(rows)


def write_range_workbook(
    path: Path,
    all_results: pd.DataFrame,
    config_rows: Sequence[Mapping[str, object]],
) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if all_results.empty:
        candidates = all_results.copy()
        confirmed = all_results.copy()
        watch = all_results.copy()
    else:
        rank = {"CONFIRMED": 0, "WATCH": 1, "REJECTED": 2}
        ordered = all_results.copy()
        ordered["_rank"] = ordered["Status"].map(rank).fillna(9)
        sort_cols = ["Actual_Date", "_rank", "Score"]
        ordered = ordered.sort_values(sort_cols, ascending=[True, True, False]).drop(columns="_rank")
        all_results = ordered
        candidates = ordered[ordered["Status"].isin(["CONFIRMED", "WATCH"])].copy()
        confirmed = ordered[ordered["Status"] == "CONFIRMED"].copy()
        watch = ordered[ordered["Status"] == "WATCH"].copy()

    perf_by_date = _performance_summary(candidates, ["Actual_Date", "Status"])
    perf_by_status = _performance_summary(candidates, ["Status"])

    if not candidates.empty:
        candidates = candidates.copy()
        candidates["Score_Band"] = (pd.to_numeric(candidates["Score"], errors="coerce") // 5 * 5).astype("Int64")
    perf_by_score = _performance_summary(candidates, ["Status", "Score_Band"]) if not candidates.empty else pd.DataFrame()

    sheets = {
        "range_candidates": candidates,
        "range_confirmed": confirmed,
        "range_watch": watch,
        "range_all_results": all_results,
        "performance_by_date": perf_by_date,
        "performance_by_status": perf_by_status,
        "performance_by_score": perf_by_score,
        "config": pd.DataFrame(config_rows),
    }

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for c in ws[1]:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")

        headers = {c.value: i for i, c in enumerate(ws[1], 1)}
        status_col = headers.get("Status")
        if status_col:
            for r in range(2, ws.max_row + 1):
                value = ws.cell(r, status_col).value
                if value in STATUS_FILLS:
                    ws.cell(r, status_col).fill = STATUS_FILLS[value]
                    ws.cell(r, status_col).font = Font(bold=True)

        for idx, cells in enumerate(ws.columns, 1):
            width = min(45, max(10, max((len(str(c.value)) if c.value is not None else 0 for c in cells), default=0) + 2))
            ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(path)
    return path
